from __future__ import annotations
import csv
import io
import sqlite3
import time
import zipfile
from collections.abc import Iterator
from pathlib import Path

import numpy as np
from tqdm import tqdm

from . import db
from .config import get_settings
from .embedding import encode_texts
from .progress import ProgressWriter
from .vector_index import VectorIndex


def normalize_row(raw: dict[str, str]) -> dict[str, any]:
    def clean(v: str | None):
        if v is None:
            return None
        v = v.strip()
        return v or None

    follower = raw.get("followerCount") or raw.get("followers") or ""
    following = raw.get("followingCount") or raw.get("following") or ""
    website = raw.get("website") or raw.get("url") or ""
    email = raw.get("email") or ""
    phone = raw.get("phone") or raw.get("phoneNumber") or ""
    bio = raw.get("bio") or raw.get("description") or ""
    return {
        "username": clean(raw.get("username")),
        "name": clean(raw.get("name")),
        "bio": clean(bio),
        "category": clean(raw.get("category")),
        "follower_count": int(follower) if follower.isdigit() else None,
        "following_count": int(following) if following.isdigit() else None,
        "website": clean(website),
        "email": clean(email),
        "phone": clean(phone),
    }


def iter_csv_rows(z: zipfile.ZipFile, member: str) -> Iterator[dict[str, str]]:
    with z.open(member) as f:
        wrapper = io.TextIOWrapper(f, encoding="utf-8", errors="replace", newline="")
        reader = csv.DictReader(wrapper)
        yield from reader


def iter_xlsx_rows(z: zipfile.ZipFile, member: str) -> Iterator[dict[str, str]]:
    from openpyxl import load_workbook

    data = z.read(member)
    bio = io.BytesIO(data)
    wb = load_workbook(bio, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        raw = {
            h: ("" if v is None else str(v)) 
            for h, v in zip(headers, values, strict=True)
        }
        yield raw


def ingest(
    zip_path: Path, 
    pattern: str, 
    dataset: str, 
    limit: int | None = None, 
    no_vectors: bool = False
):
    settings = get_settings()
    conn = db.connect(settings.db_path)
    status = ProgressWriter(Path("data/status.json"))
    index = VectorIndex.load(settings.index_dir)
    
    with zipfile.ZipFile(zip_path) as z:
        members = [m for m in z.namelist() if Path(m).match(pattern)]
        if not members:
            raise SystemExit(f"No members match pattern {pattern}")
        
        for member in members:
            # skip if processed
            cur = conn.execute(
                "SELECT 1 FROM processed_files WHERE file_name=?", (member,)
            ).fetchone()
            if cur:
                continue
            
            rows_buffer = []
            text_concat: list[str] = []
            batch_size = settings.batch_size
            total_rows = 0
            start_time = time.time()
            
            # Choose iterator based on file extension
            if member.lower().endswith(".xlsx"):
                iterator = iter_xlsx_rows(z, member) 
            else:
                iterator = iter_csv_rows(z, member)
            
            pbar = tqdm(desc=f"{member}", unit="rows")
            
            for raw in iterator:
                norm = normalize_row(raw)
                total_rows += 1
                merged_text = " | ".join(
                    filter(
                        None,
                        [
                            norm.get("username"),
                            norm.get("name"),
                            norm.get("bio"),
                            norm.get("category"),
                            norm.get("website"),
                            norm.get("email"),
                            norm.get("phone"),
                        ],
                    )
                )
                norm["text_concat"] = merged_text
                rows_buffer.append(
                    (
                        dataset,
                        norm["username"],
                        norm["name"],
                        norm["bio"],
                        norm["category"],
                        norm["follower_count"],
                        norm["following_count"],
                        norm["website"],
                        norm["email"],
                        norm["phone"],
                        norm["text_concat"],
                    )
                )
                text_concat.append(merged_text)
                
                if len(rows_buffer) >= batch_size:
                    flush_batch(conn, rows_buffer, text_concat, index, no_vectors)
                    rows_buffer.clear()
                    text_concat.clear()
                    status.update(
                        {
                            "dataset": dataset,
                            "current_file": member,
                            "rows_processed": total_rows,
                            "files_completed": len(members),
                            "batch_time_sec": round(time.time() - start_time, 2),
                        }
                    )
                    if limit and total_rows >= limit:
                        break
                pbar.update(1)
            
            # Final partial batch
            if rows_buffer:
                flush_batch(conn, rows_buffer, text_concat, index, no_vectors)
            
            with conn:
                conn.execute(
                    """INSERT OR REPLACE INTO processed_files(
                        dataset, file_name, row_count, completed_at
                    ) VALUES (?,?,?,datetime('now'))""",
                    (dataset, member, total_rows),
                )
            
            status.update(
                {
                    "dataset": dataset, 
                    "current_file": None, 
                    "rows_processed": total_rows, 
                    "file_completed": member
                }, 
                force=True
            )
            
            if limit and total_rows >= limit:
                break
    
    # save index at end
    if index and not no_vectors:
        index.save(settings.index_dir)


def flush_batch(
    conn: sqlite3.Connection,
    rows_buffer,
    text_concat,
    index: VectorIndex | None,
    no_vectors: bool,
):
    """Insert batch of rows into DB and FTS table."""
    with db.transaction(conn):
        db.bulk_insert_leads(conn, rows_buffer)
    
    # Insert into FTS - need last rowids to link correctly
    cur = conn.execute("SELECT MAX(id) FROM leads")
    max_id = cur.fetchone()[0]
    count = len(rows_buffer)
    first_id = max_id - count + 1
    
    # Insert FTS content with proper rowids
    fts_rows = []
    for _, row in enumerate(rows_buffer):
        fts_rows.append((row[1], row[2], row[3], row[4], row[7], row[8], row[9]))
    
    with db.transaction(conn):
        conn.executemany(
            """INSERT INTO leads_fts(
                rowid, username, name, bio, category, website, email, phone
            ) VALUES (?,?,?,?,?,?,?,?)""",
            [(first_id + i, *fr) for i, fr in enumerate(fts_rows)],
        )

    # Handle vector embeddings if enabled
    if not no_vectors and index:
        embeddings = encode_texts(text_concat)
        if index is None:
            index = VectorIndex(embeddings.shape[1])
        ids = np.arange(first_id, first_id + len(rows_buffer))
        index.add(embeddings, ids)
        
        # Periodic flush based on settings
        settings = get_settings()
        if settings.flush_every <= 1:
            index.save(settings.index_dir)
